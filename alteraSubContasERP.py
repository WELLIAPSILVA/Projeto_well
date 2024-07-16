#Altera subcontas e unidade no sistema flex RPInfo
import openpyxl
import pyautogui
import time
import keyboard

usuario=pyautogui.prompt(text='digite seu usuário',title='Usuário')
senha= pyautogui.password(text='Digite sua senha',title='senha',mask='*')
URL_arquivo=pyautogui.prompt(text='digite o local e nome do arquivo',title='Local do arquivo')
unidade=pyautogui.prompt(text='digite a unidade com a seguinda mascara 000',title='Unidade')
subcontas=pyautogui.prompt(text='digite o subcontas com a seguinte mascara 000',title='Subcontas')
time.sleep(1)
# Abre o programa ERP.exe
pyautogui.hotkey('winleft', 'r') # Pressiona a tecla Windows + Esquerda para abrir o menu Iniciar
pyautogui.typewrite('C:\\erp\\erp.exe')  # Digita o nome do programa
pyautogui.press('enter')  # Pressiona Enter para abrir o programa

pyautogui.press('enter')  # Pressiona Enter para fazer login
# Abre o programa ERP.exe
pyautogui.press('enter')

time.sleep(10)
# Faz login com usuário e senha
pyautogui.typewrite(usuario)  # Digita o usuário
pyautogui.press('enter')  #  Pressiona Enter
pyautogui.typewrite(senha)  # Digita a senha
pyautogui.press('enter')  # Pressiona Enter para fazer login

# Aguarda um pouco para que o login seja realizado
time.sleep(5)

# Navega até o menu desejado e realiza a ação
pyautogui.click(533, 30)  # Clica no menu desejado
time.sleep(1)
pyautogui.click(600, 185)  # Clica no item de menu desejado

time.sleep(5)
# Abrir o arquivo Excel
wb = openpyxl.load_workbook(URL_arquivo)
sheet = wb['Planilha1']  # Substituir por nome da planilha

# Posições do mouse para clique e colagem
posicao_campo_especifico = (556, 264)  # Substituir pelascoordenadas do campo específico
posicao_botao_alterar = (1392, 264)  # Substituir pelas coordenadas do botão Alterar
#posicao_lista_selecao1 = (861, 635)  # Substituir pelas coordenadas da lista de seleção 1
#posicao_lista_selecao2 = (834, 507)  # Substituir pelas coordenadas da lista de seleção 2

# Loop pelas células da coluna Transação
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    
    
    import pyperclip
    

    for cell in row:
        codigo_transacao = cell.value
        print(codigo_transacao)

    # Copiar o código de transação como texto
        pyperclip.copy(str(codigo_transacao))

    # Mover o mouse para o campo específico e colar o código
        pyautogui.moveTo(posicao_campo_especifico)
        pyautogui.click()
        
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('enter')
        
        # Mover o mouse para o botão Alterar e clicar
        pyautogui.moveTo(posicao_botao_alterar)
        pyautogui.click()
        time.sleep(1)

        # Mover o mouse para a lista de seleção 1 e selecionar
        #pyautogui.moveTo(posicao_lista_selecao1)
        #pyautogui.click()
        #time.sleep(1)
        pyautogui.typewrite(str(subcontas))#subcontas
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.press('enter')
        
        # Mover o mouse para a lista de seleção 2 e selecionar
        #pyautogui.moveTo(posicao_lista_selecao2)
        #pyautogui.click()
        #time.sleep(1)
        pyautogui.typewrite(str(unidade))#unidade
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.press('enter')
        
             # Verificar se a tecla 'esc' foi pressionada
        if keyboard.is_pressed('esc'):
            print("Execução pausada pelo usuário.")
            break        
        # Adicionar um delay para evitar que as ações sejam realizadas muito rapidamente
        time.sleep(1)
